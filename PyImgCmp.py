import os
from PIL import Image

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.styles import Color, PatternFill

import collections
image_fields = ['path', 'type', 'width', 'height', 'mode']

class ImageInfo(collections.namedtuple('ImageInfo', image_fields)):

    def to_str_row(self):
        return ("%s\t%d\t%d\t%s\t%s" % (
            self.mode,
            self.width,
            self.height,
            self.type,
            self.path.replace('\t', '\\t'),
        ))

    def to_str_row_verbose(self):
        return ("%s\t%d\t%d\t%s\t%s\t##%s" % (
            self.mode,
            self.width,
            self.height,
            self.type,
            self.path.replace('\t', '\\t'),
            self))

class FileCompareResult(object):
    CODE_ERROR = -1
    CODE_WARN = -2
    CODE_SUCCESS = 0

    def __init__(self, code):
        self.code = code

class ImageCompareResult(FileCompareResult):
    def __init__(self, imagefile, otherfile, diff, image: ImageInfo, code = FileCompareResult.CODE_SUCCESS, reason = ''):
        super().__init__(code)
        self.imagefile = imagefile
        self.otherfile = otherfile
        self.image = image
        self.diff = diff
        self.reason = reason

    def __eq__(self, other) -> bool:
        return self.imagefile == other.imagefile and self.otherfile == other.otherfile

    def __str__(self):
        return '{},{},{:.10f},f:{},w:{:d},h:{:d},m:{},{}'.format(self.imagefile, self.otherfile, self.diff, self.image.type, self.image.width, self.image.height, self.image.mode, self.reason)

    def iterator(self):
        return str(self).split(',')

    def correct(self) -> bool:
        split = self.imagefile.split('\\')
        imagefile_name = split[len(split) - 1]
        split = self.otherfile.split('\\')
        otherfile_name = split[len(split) - 1]
        return imagefile_name == otherfile_name

class ImageComparater(object):
    IGNORE_DIFF_SCORE = -1
    EQUAL_DIFF_SCORE = 0

    def calculate_diff(self, imagefile1, imagefile2) -> ImageCompareResult:
        return ImageCompareResult('', '', ImageComparater.EQUAL_DIFF_SCORE, [], FileCompareResult.CODE_SUCCESS)

class ImageSimpleComparater(ImageComparater):
    def calculate_diff(self, imagefile1, imagefile2):
        img1 = Image.open(imagefile1)
        img2 = Image.open(imagefile2)
        imginfo1 = ImageInfo(path = imagefile1, type = img1.format, width = img1.width, height = img1.height, mode = img1.mode)
        imginfo2 = ImageInfo(path = imagefile2, type = img2.format, width = img2.width, height = img2.height, mode = img2.mode)
        img1.close()
        img2.close()
        if not imginfo1.type == imginfo2.type or not imginfo1.width == imginfo2.width or not imginfo1.height == imginfo2.height:
            return ImageCompareResult(imagefile1, imagefile2, ImageComparater.IGNORE_DIFF_SCORE, imginfo1, FileCompareResult.CODE_WARN, 'simple image info not match')
        splits = imagefile1.split('\\')
        filename1 = splits[len(splits) - 1]
        splits = imagefile2.split('\\')
        filename2 = splits[len(splits) - 1]
        if filename1 == filename2:
            return ImageCompareResult(imagefile1, imagefile2, ImageComparater.EQUAL_DIFF_SCORE, imginfo1)
        else:
            return ImageCompareResult(imagefile1, imagefile2, ImageComparater.IGNORE_DIFF_SCORE, imginfo1, FileCompareResult.CODE_ERROR, 'filename not match')

class ImageFileIterator(object):
    def __init__(self, folders):
        for folder in folders:
            if not os.path.isdir(folder):
                raise ValueError('Invalid input folder {} when init'.format(folder))
        self.folders = folders

    def is_imagefile(self, imagepath):
        if not os.path.isfile(imagepath):
            return False
        split = imagepath.split('.')
        if 'png' == split[len(split) - 1] or 'jpg' == split[len(split) - 1]:
            return True
        return False

    def process(self, file):
        pass

    def iterator(self, folder):
        if not os.path.isdir(folder):
            raise ValueError('Invalid input folder {} when iterator'.format(folder))
        for file in os.listdir(folder):
            file = os.path.join(folder, file)
            if os.path.isdir(file):
                self.iterator(file)
            elif self.is_imagefile(file):
                self.process(file)

    def start(self):
        for folder in self.folders:
            self.iterator(folder)

class ImageFileCompareTask(ImageFileIterator):
    def __init__(self, imagefile, folders):
        super().__init__(folders)
        self.imagefile = imagefile
        self.comparater = ImageSimpleComparater()
        self.equal_list = []
        self.similar_list = []
        self.warn_list = []
        self.error_list = []

    def insert_result(self, result: ImageCompareResult):
        if FileCompareResult.CODE_SUCCESS == result.code:
            self.equal_list.append(result)
        elif FileCompareResult.CODE_WARN == result.code:
            self.warn_list.append(result)
        elif FileCompareResult.CODE_ERROR == result.code:
            self.error_list.append(result)

    def process(self, imagefile):
        result = self.comparater.calculate_diff(self.imagefile, imagefile)
        self.insert_result(result)

    def start(self):
        super().start()
        len_equal = len(self.equal_list)
        len_error = len(self.error_list)
        len_warn = len(self.warn_list)
        print('{}: equal {:d}, warn {:d}, error {:d}'.format(self.imagefile, len_equal, len_warn, len_error))
        if not len_equal == 0:
            self.warn_list.clear
            self.error_list.clear
        elif not len_error == 0:
            self.warn_list.clear
        elif not len_warn == 0:
            self.error_list.append(self.warn_list[0])
            self.warn_list.clear
        else:
            print('Can\'t get here')

        if len_error > 1:
            del self.error_list[1:len(self.error_list)]

class ResultPrinter(object):
    def print(self, compare_task: ImageFileCompareTask):
        pass

class ExcelPrinter(ResultPrinter):
    EQUAL_SHEET_NAME = 'equal'
    SIMILAR_SHEET_NAME = 'similar'
    ERROR_SHEET_NAME = 'error'

    def openxlsx(self, filepath) -> Workbook:
        folder = '.'
        filename = filepath
        if '\\' in filepath:
            folder = filepath[:filepath.rindex('\\')]
            filename = filepath[filepath.rindex('\\') + 1:]
        
        for file in os.listdir(folder):
            if os.path.isfile(file):
                if filename == file:
                    return load_workbook(filepath)
        wbk = Workbook()
        wbk.save(filepath)
        return load_workbook(filepath)

    def __init__(self, filepath, default_sheet_name = EQUAL_SHEET_NAME):
        self.filepath = filepath
        self.xlsxfile = self.openxlsx(filepath)
        ws = self.xlsxfile.active
        ws.title = default_sheet_name
        self.xlsxfile.save(filepath)
        self.sheet_dict = dict()
        self.sheet_dict[default_sheet_name] = 0
        self.fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('C4C4C4'))

    def finish_print(self):
        for ws in self.xlsxfile.worksheets:
	        #format column width
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
        self.xlsxfile.save(self.filepath)

    def print_sheet(self, sheet_name, results_list):
        if not sheet_name in self.xlsxfile:
            self.xlsxfile.create_sheet(sheet_name)
        ws = self.xlsxfile[sheet_name]
        for result in results_list:
            ws.append(result.iterator())
            if not result.correct():
                self.sheet_dict[sheet_name] = self.sheet_dict.get(sheet_name, 0) + 1
                for row in ws.iter_rows(min_row=ws.max_row):
                    for cell in row:
                        cell.fill = self.fill
                ws.cell(row = 1, column = 8, value = int(self.sheet_dict[sheet_name]))

    def print(self, compare_task: ImageFileCompareTask):
        equal_list = compare_task.equal_list
        similar_list = compare_task.similar_list
        error_list = compare_task.error_list
        len_equal = len(equal_list)
        len_similar = len(similar_list)
        len_error = len(error_list)
        if not len_equal == 0:
            #print equal sheet
            #print('{} equal record {}'.format(compare_task.imagefile, len_equal))
            self.print_sheet(ExcelPrinter.EQUAL_SHEET_NAME, equal_list)
        elif not len_similar == 0:
            #print similar sheet
            #print('{} similar record {}'.format(compare_task.imagefile, len_similar))
            self.print_sheet(ExcelPrinter.SIMILAR_SHEET_NAME, similar_list)
        elif not len_error == 0:
            #print error sheet
            #print('{} error record {}'.format(compare_task.imagefile, len_error))
            self.print_sheet(ExcelPrinter.ERROR_SHEET_NAME, error_list)
        else:
            #no result
            print('{} no record !!!'.format(compare_task.imagefile))
        self.finish_print()

class ImageFolderCompareTask(ImageFileIterator):
    def __init__(self, folders1, folders2, printer: ResultPrinter):
        super().__init__(folders1)
        self.otherfolders = folders2
        self.printer = printer

    def start_file_compare_task(self, file) -> ImageFileCompareTask:
        file_compare_task = ImageFileCompareTask(file, self.otherfolders)
        file_compare_task.start()
        return file_compare_task

    def process(self, file):
        file_compare_task = self.start_file_compare_task(file)
        # print results
        self.printer.print(file_compare_task)

if __name__ == '__main__':
    ui_folder = ['setting']
    proj_folder = ['future-skin-blue']
    xlsx_filepath = 'setting.xlsx'

    comparater = ImageFolderCompareTask(proj_folder, ui_folder, ExcelPrinter(xlsx_filepath))
    comparater.start()


