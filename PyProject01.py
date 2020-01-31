import os
from PIL import Image
from imagehash import ImageHash
import imagehash
import cv2

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from skimage.measure import compare_ssim

ui_folder = 'gray'
project_folder = 'blue'
output_xlsx = 'result_gb.xlsx'
max_record = 10

RESULT_CODE_ERROR = -1

class ImageComparater(object):

    def compare_image(self, imagepath1, imagepath2):
        pass

class ImageHashComparater(ImageComparater):

    def average_hash(self, image):
        image = image.convert("L")
        pixels = numpy.asarray(image)
        avg = pixels.mean()
        diff = pixels > avg
        return ImageHash(diff)

    def compare_image(self, imagepath1, imagepath2):
        img1 = Image.open(imagepath1)
        img2 = Image.open(imagepath2)
        '''
        return imagehash.dhash(img1) - imagehash.dhash(img2)
        return imagehash.average_hash(img1) - imagehash.average_hash(img2)
        return imagehash.phash(img1) - imagehash.phash(img2)
        return imagehash.whash(img1) - imagehash.whash(img2)
        '''
        return imagehash.dhash_vertical(img1) - imagehash.dhash_vertical(img2)

class ImageHistComparater(ImageComparater):

    def compare_image(self, imagepath1, imagepath2):
        img1 = cv2.imread(imagepath1, cv2.IMREAD_UNCHANGED)
        img2 = cv2.imread(imagepath2, cv2.IMREAD_UNCHANGED)

        imgHist1 = cv2.calcHist([img1], [0], None, [256], [0, 256])
        imgHist2 = cv2.calcHist([img2], [0], None, [256], [0, 256])

        img_hist_diff = cv2.compareHist(imgHist1, imgHist2, cv2.HISTCMP_BHATTACHARYYA)
        #print 'img_hist_diff=', img_hist_diff
        img_template_match = cv2.matchTemplate(imgHist1, imgHist2, cv2.TM_CCOEFF_NORMED)[0][0]
        img_template_diff = 1 - img_template_match
        #print 'img_template_diff=', img_template_diff
        return (img_hist_diff / 10) + img_template_diff

class ImageSSIMComparater(ImageComparater):

    def compare_image(self, imagepath1, imagepath2):
        img1 = cv2.imread(imagepath1, cv2.IMREAD_UNCHANGED)
        img2 = cv2.imread(imagepath2, cv2.IMREAD_UNCHANGED)
        if not img1.dtype == img2.dtype or not img1.shape == img2.shape:
            #print ('Invalid compare inputs: left={}{}, right={}{}'.format(img1.dtype, img1.shape, img2.dtype, img2.shape))
            return RESULT_CODE_ERROR;

        #img1 = img1[:,:,3]
        #img2 = img2[:,:,3]
        
        try:
            bgrScore = compare_ssim(img1, img2, multichannel=True)
        except ValueError:
            print ('Invalid compare inputs: left={}{}{}, right={}{}{}'.format(imagepath1, img1.dtype, img1.shape, imagepath2, img2.dtype, img2.shape))
            return RESULT_CODE_ERROR;
        return 1.0 - bgrScore

class CmpResult(object):
    def __init__(self, leftpath, rightpath, diff):
        self.leftpath = leftpath
        self.rightpath = rightpath
        self.diff = diff

    def __str__(self):
        return '{0:<15},{0:<15},{1:<.10f}'.format(self.leftpath, self.rightpath, self.diff)

class ImageFileFolderComparater(object):

    def __init__(self, max_record, printer, comparater):
        self.max_record = max_record
        self.result_list = []
        self.printer = printer
        self.comparater = comparater

    def is_imagefile(self, imagepath):
        if not os.path.isfile(imagepath):
            return False
        split = imagepath.split('.')
        if 'png' == split[len(split) - 1] or 'jpg' == split[len(split) - 1]:
            return True
        return False

    def add_result(self, result): 
        length = len(self.result_list)
        position = -1
        for i in range(length):
            record = self.result_list[i]
            if record.diff > result.diff:
                position = i
                break
        if length < self.max_record:
            self.result_list.append(0)
            if -1 == position:
                self.result_list[length] = result
                return
            length += 1
        else:
            if -1 == position:
                return
        for i in range(length - 1, position, -1):
            self.result_list[i] = self.result_list[i - 1]
        self.result_list[position] = result

    def print_results(self):
        for result in self.result_list:
            print (result.leftpath, ' - ', result.rightpath, '=', result.diff)

    def compare_file_folder(self, imagepath, folderpath):
        for file in os.listdir(folderpath):
            file = os.path.join(folderpath, file)
            if self.is_imagefile(file):
                diff = self.comparater.compare_image(imagepath, file)
                if not RESULT_CODE_ERROR == diff:
                    #print ('compare_image: {} with {}, diff={}'.format(imagepath, file, diff))
                    self.add_result(CmpResult(imagepath, file, diff))
                #self.print_results()
            elif os.path.isdir(file):
                self.compare_file_folder(imagepath, file)

    def compare(self, leftfolder, rightfolder):
        for file in os.listdir(leftfolder):
            file = os.path.join(leftfolder, file)
            if self.is_imagefile(file):
                del self.result_list[:]
                self.compare_file_folder(file, rightfolder)
                printer.printxlsx(self.result_list)
            elif os.path.isdir(file):
                self.compare(file, rightfolder)
        

class XlsxPrinter(object):

    def __init__(self, filename):
        self.filename = filename
        self.current_row = 1

    def openxlsx(self, filename):
        for file in os.listdir('.'):
            if os.path.isfile(file):
                if filename == file:
                    return load_workbook(filename)
        wbk = Workbook()
        ws = wbk.active
        ws.title = filename
        wbk.save(filename)
        return load_workbook(filename)

    def begin_print(self):
        self.xlsx_file = self.openxlsx(self.filename)

    def printxlsx(self, compare_results):
        ws = self.xlsx_file.active
        length = len(compare_results)
        for i in range(0, length):
            result = compare_results[i]
            ws.cell(row = self.current_row, column = 1, value = result.leftpath)
            ws.cell(row = self.current_row, column = 2, value = result.rightpath)
            ws.cell(row = self.current_row, column = 3, value = '{:<.10f}'.format(result.diff))
            self.current_row += 1

        self.current_row += 1
        self.finish_print()

    def finish_print(self):
        ws = self.xlsx_file.active
        #format column width
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        self.xlsx_file.save(self.filename)    

if __name__ == '__main__':
    printer = XlsxPrinter(output_xlsx)    
    printer.begin_print()
    comparater = ImageFileFolderComparater(max_record, printer, ImageSSIMComparater())
    comparater.compare(project_folder, ui_folder)
    printer.finish_print()